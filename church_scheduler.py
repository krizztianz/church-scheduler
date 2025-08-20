
import argparse
import os
import sys
from datetime import date, timedelta
from collections import deque, defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


def info(msg: str):
    print(f"[INFO] {msg}", file=sys.stdout)

def success(msg: str):
    print(f"[OK] {msg}", file=sys.stdout)

def error(msg: str):
    print(f"[ERROR] {msg}", file=sys.stderr)


# -------------------------
# Helpers
# -------------------------
def sundays_in_month(year: int, month: int):
    d = date(year, month, 1)
    # move to first Sunday
    while d.weekday() != 6:  # Monday=0 ... Sunday=6
        d += timedelta(days=1)
    out = []
    while d.month == month:
        out.append(d)
        d += timedelta(days=7)
    return out

def read_master(path: str):
    df = pd.read_excel(path)
    # detect header row by finding the row where column 'Tugas' equals the first task label
    # In provided sample, row 0 stores task labels across columns (Tugas, Unnamed:3...)
    # We'll treat the first row as horizontal header for task names
    header_row = df.iloc[0]
    task_cols = df.columns[2:11]  # from 'Tugas' onward until 'Multimedia'
    # Map real human-readable task names from first row
    col_name_map = {}
    for c in task_cols:
        label = str(header_row[c]).strip()
        if label == "nan":
            continue
        col_name_map[c] = label

    # Slice data rows (skip first row that contains labels)
    data = df.iloc[1:].copy()
    # Normalize name
    data['Nama'] = data['Nama'].astype(str).str.strip()
    # Normalize Penatua flag (x/X/1/True)
    def to_bool(v):
        if pd.isna(v): return False
        s = str(v).strip().lower()
        return s in ('x','1','true','ya','yes','y')
    data['Penatua'] = data['Penatua'].apply(to_bool)

    # Build per-task ability flags from mapped columns
    abilities = {}
    for raw_col, task in col_name_map.items():
        abilities[task] = data[raw_col].apply(to_bool).tolist()

    # Build records list
    records = []
    for i, row in data.reset_index(drop=True).iterrows():
        rec = {'Nama': row['Nama'], 'Penatua': bool(row['Penatua'])}
        for raw_col, task in col_name_map.items():
            rec[task] = bool(abilities[task][i])
        records.append(rec)

    people = pd.DataFrame(records)
    return people, list(col_name_map.values())

def make_rotation(candidates):
    """Return a deque rotation for fairness."""
    return deque(candidates)

def pick_from_rotation(rot, k=1, exclude=set()):
    picked = []
    tried = 0
    n = len(rot)
    if n == 0 or k == 0:
        return picked
    while len(picked) < k and tried < n*2:  # tolerate cycling twice if necessary
        person = rot[0]
        rot.rotate(-1)
        tried += 1
        if person in exclude:
            continue
        if person not in picked:
            picked.append(person)
    return picked

def build_scheduler(people, seed_key):
    """Create rotations per category (Penatua vs Jemaat) and per task ability."""
    import random
    random.seed(seed_key)

    elders = people[people['Penatua'] == True]['Nama'].tolist()
    members = people[people['Penatua'] == False]['Nama'].tolist()

    rotations = {'ALL': make_rotation(sorted(people['Nama'].tolist())),
                 'ELDER': make_rotation(sorted(elders)),
                 'MEMBER': make_rotation(sorted(members))}

    # per-task rotations (ability-aware)
    task_rot = {}
    for task in [c for c in people.columns if c not in ('Nama','Penatua')]:
        able = people[people[task] == True]
        task_rot[task] = {
            'ELDER': make_rotation(sorted(able[able['Penatua'] == True]['Nama'].tolist())),
            'MEMBER': make_rotation(sorted(able[able['Penatua'] == False]['Nama'].tolist())),
            'ALL': make_rotation(sorted(able['Nama'].tolist()))
        }
    return rotations, task_rot

def schedule_month(people, year, month, prefer_no_repeat_non_elder=True, pjemaat_count=3):
    """Generate schedule with constraint: a person should not receive multiple tasks in the same week.
    We still guarantee task ability and category rules. If pool is insufficient, we relax
    the no-repeat constraint as a *last resort* to fill slots.
    """
    # Business rules per prompt

    NEEDS = {
        'DP/PA':      {'count':1, 'who':'ELDER'},
        'W/PB':       {'count':1, 'who':'ELDER'},
        'Persembahan':{'count':1, 'who':'ELDER'},
        'Kolektan':   {'count':1, 'who':'ELDER'},
        'P. Jemaat':  {'count':pjemaat_count, 'who':'MIXED'},  # 3 or 4, we'll make 3 by default
        'Lektor':     {'count':2, 'who':'MEMBER'},
        'Pemusik':    {'count':2, 'who':'MEMBER'},
        'Multimedia': {'count':1, 'who':'ALL'},  # no explicit restriction
        'Prokantor':  {'count':2, 'who':'MEMBER'},
    }

    sundays = sundays_in_month(year, month)
    rotations, task_rot = build_scheduler(people, seed_key=f"{year}-{month}")

    def assign_candidates(task, who, need, hard_exclude=set(), soft_member_exclude=set()):
        assigned = []

        def pick(group_key, k, exclude=set()):
            return pick_from_rotation(task_rot[task][group_key], k=k, exclude=exclude)

        if who == 'ELDER':
            # Strictly avoid weekly duplicates
            assigned = pick('ELDER', k=need, exclude=hard_exclude)
            if len(assigned) < need:
                # If not enough elders available this week, allow weekly repeats only as last resort
                more = pick('ELDER', k=need-len(assigned))
                assigned += [m for m in more if m not in assigned]
            return assigned[:need]

        if who == 'MEMBER':
            # First avoid weekly duplicates + avoid last-week repeats
            ex = set(hard_exclude) | (soft_member_exclude if prefer_no_repeat_non_elder else set())
            assigned = pick('MEMBER', k=need, exclude=ex)
            if len(assigned) < need:
                # Relax last-week constraint but keep weekly uniqueness
                ex2 = set(hard_exclude)
                more = pick('MEMBER', k=need-len(assigned), exclude=ex2)
                for m in more:
                    if m not in assigned:
                        assigned.append(m)
            if len(assigned) < need:
                # As a last resort, allow weekly repeat
                more = pick('MEMBER', k=need-len(assigned))
                for m in more:
                    if m not in assigned:
                        assigned.append(m)
            return assigned[:need]

        if who == 'ALL':
            assigned = pick('ALL', k=need, exclude=hard_exclude)
            if len(assigned) < need:
                more = pick('ALL', k=need-len(assigned))
                assigned += [m for m in more if m not in assigned]
            return assigned[:need]

        if who == 'MIXED':
            # Aim for 2 members + 1 elder (for need=3) while avoiding weekly dupes
            mem_need = min(2, need)
            eld_need = need - mem_need

            # 1) Members first, avoid weekly dupes + last-week repeats if configured
            ex_mem = set(hard_exclude) | (soft_member_exclude if prefer_no_repeat_non_elder else set())
            members = pick('MEMBER', k=mem_need, exclude=ex_mem)
            if len(members) < mem_need:
                # Relax last-week rule but keep weekly uniqueness
                members += [m for m in pick('MEMBER', k=mem_need-len(members), exclude=set(hard_exclude)) if m not in members]
            # 2) Elders, avoid weekly dupes
            elders = pick('ELDER', k=eld_need, exclude=hard_exclude)

            combined = members + elders

            # If shortage, try fill from ALL (still avoid weekly dupes first)
            if len(combined) < need:
                more = pick('ALL', k=need-len(combined), exclude=hard_exclude)
                for m in more:
                    if m not in combined:
                        combined.append(m)

            # Absolute fallback: allow weekly repeat as very last resort
            if len(combined) < need:
                more = pick('ALL', k=need-len(combined))
                for m in more:
                    if m not in combined:
                        combined.append(m)

            return combined[:need]

    # track recently assigned for MEMBERS only if prefer_no_repeat_non_elder
    last_week_assigned_member = set()  # for avoiding back-to-back member repeats
    weekly_assigned_all = set()  # for avoiding multi-task in the same week

    results = defaultdict(dict)
    tasks_in_master = [c for c in people.columns if c not in ('Nama','Penatua')]
    for day in sundays:
        weekly_assigned_all = set()  # reset for each Sunday
        weekly_assigned_member = set()  # to avoid duplicates within same week
        # iterate over needed tasks if present in master
        for task, spec in NEEDS.items():
            if task not in tasks_in_master:
                continue

            assigned = assign_candidates(
                task=task,
                who=spec['who'],
                need=spec['count'],
                hard_exclude=weekly_assigned_all,
                soft_member_exclude=last_week_assigned_member
            )

            # Update tracking
            for name in assigned:
                weekly_assigned_all.add(name)
                if name in set(people[people['Penatua']==False]['Nama']):
                    weekly_assigned_member.add(name)

            results[day][task] = assigned

        # end tasks loop
        last_week_assigned_member = weekly_assigned_member

    return sundays, results

def write_excel(output_path, sundays, schedule):
    wb = Workbook()
    ws = wb.active
    ws.title = "Jadwal Bulanan"

    # header row
    ws.cell(row=1, column=1, value="WAKTU").font = Font(bold=True)
    for idx, d in enumerate(sundays, start=2):
        ws.cell(row=1, column=idx, value=f"MINGGU, {d.day:02d} {d.strftime('%B %Y').upper()}\nPkl. 07.00 Wib,").alignment = Alignment(wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = 28

    # task order following Output-Example order
    desired_order = ["DP/PA","W/PB","Persembahan","Kolektan","P. Jemaat","Lektor","Prokantor","Pemusik","Multimedia"]
    tasks_present = [t for t in desired_order if any(t in schedule[d] for d in sundays)]

    # styling
    th_fill = PatternFill("solid", fgColor="BDD7EE")
    thin = Side(style="thin", color="000000")

    ws.column_dimensions['A'].width = 18

    # write task rows
    for r, task in enumerate(tasks_present, start=2):
        ws.cell(row=r, column=1, value=task).font = Font(bold=True)
        for c, d in enumerate(sundays, start=2):
            names = schedule[d].get(task, [])
            text = "\n".join(names) if isinstance(names, list) else str(names)
            ws.cell(row=r, column=c, value=text).alignment = Alignment(wrap_text=True, vertical="top")
            # borders
            for cc in range(1, len(sundays)+2):
                ws.cell(row=r, column=cc).border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # header style & borders
    for c in range(1, len(sundays)+2):
        cell = ws.cell(row=1, column=c)
        cell.fill = th_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, horizontal="center")

        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb.save(output_path)

def main():
    ap = argparse.ArgumentParser(description="Generate jadwal ibadah bulanan from Master.xlsx")
    ap.add_argument("--pjemaat-count", type=int, default=3,
                    help="Jumlah P. Jemaat per minggu (default=3, max=4)")
    ap.add_argument("--master", required=True, help="Path to Master.xlsx")
    ap.add_argument("--year", type=int, required=True)
    ap.add_argument("--month", type=int, required=True)
    ap.add_argument("--output", default="output/Jadwal-Bulanan.xlsx")
    ap.add_argument("--repeat-non-elder", action="store_true",
                    help="Allow non-elders to repeat on consecutive weeks when needed (default: tries to avoid repeats if pool is sufficient)")
    ap.add_argument("--verbose", action="store_true",
                    help="Print extra info about assignments and rotations")
    args = ap.parse_args()

    try:
        if not os.path.exists(args.master):
            error(f"Master file not found: {args.master}")
            return

        people, tasks = read_master(args.master)
        if people.empty:
            error("Master.xlsx contains no usable data.")
            return

        sundays, sched = schedule_month(
            people, args.year, args.month,
            prefer_no_repeat_non_elder=not args.repeat_non_elder,
            pjemaat_count=min(max(args.pjemaat_count,1),4)
        )

        write_excel(args.output, sundays, sched)
        success(f"Generated {args.output} for {len(sundays)} Sundays.")

        if args.verbose:
            info(f"Total people: {len(people)} | Tasks detected: {tasks}")
            for d in sundays:
                info(f"{d}: " + ", ".join(f"{t}={len(sched[d][t])}" for t in sched[d]))

    except Exception as e:
        error(f"Unexpected error: {e}")

if __name__ == "__main__":
    main()
